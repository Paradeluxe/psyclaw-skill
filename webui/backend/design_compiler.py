"""Compile Builder design.json -> pure PsychoPy script.

Draw loop mirrors PsychoPy Builder semantics:

  * per-routine Clock + continueRoutine flag
  * component lifecycle NOT_STARTED -> STARTED -> FINISHED
  * keyboard: clearEvents on onset, timeStamped getKeys, force_end
  * fixed duration ends on clock; open-ended (-1) waits for force_end / escape
  * flow loops expand with thisN / loop name; optional conditions (stimlist)
  * conditionsFile or embedded conditions[] — nReps × rows; $var in params
"""
from __future__ import annotations

import csv
import json
import random
from pathlib import Path
from typing import Any, Dict, List, Optional

from compiler import HEADER, compile_spec

_RUNNER = r'''
import json
import random
import re
import csv as _csv
import shutil as _shutil
from pathlib import Path as _Path

# --- inlined trial_metrics.py (corrAns/corr + run summary) ---
__TRIAL_METRICS_SRC__
# --- end trial_metrics ---

DESIGN = json.loads(__DESIGN_JSON__)
# Session from mission-control Run form — never PsychoPy gui.Dlg / expInfo
SESSION = json.loads(__SESSION_JSON__)

disp = DESIGN.get("display") or {}
size = disp.get("size") or [1024, 768]
full = bool(disp.get("fullscreen", True))
try:
    _scr_idx = int(disp.get("screen", 0) or 0)
except (TypeError, ValueError):
    _scr_idx = 0
if _scr_idx < 0:
    _scr_idx = 0

def _bgcolor_of(v):
    """design.display.bgcolor hex/name/rgb → PsychoPy color [-1,1]."""
    if isinstance(v, (list, tuple)) and len(v) >= 3:
        try:
            return [float(v[0]), float(v[1]), float(v[2])]
        except (TypeError, ValueError):
            return [-1, -1, -1]
    s = str(v or "#000000").strip().lower()
    if s.startswith("$"):
        # Window color is fixed at start; $stimlist vars cannot resolve here.
        return [-1, -1, -1]
    named = {
        "black": [-1, -1, -1], "white": [1, 1, 1],
        "gray": [0, 0, 0], "grey": [0, 0, 0],
        "darkgray": [-0.5, -0.5, -0.5], "darkgrey": [-0.5, -0.5, -0.5],
        "lightgray": [0.5, 0.5, 0.5], "lightgrey": [0.5, 0.5, 0.5],
        "red": [1, -1, -1], "green": [-1, 1, -1], "blue": [-1, -1, 1],
        "yellow": [1, 1, -1], "cyan": [-1, 1, 1], "magenta": [1, -1, 1],
    }
    if s in named:
        return named[s]
    if s.startswith("#"):
        hx = s[1:]
        if len(hx) == 3:
            hx = "".join(ch * 2 for ch in hx)
        if len(hx) == 6:
            try:
                r = int(hx[0:2], 16) / 127.5 - 1.0
                g = int(hx[2:4], 16) / 127.5 - 1.0
                b = int(hx[4:6], 16) / 127.5 - 1.0
                return [max(-1.0, min(1.0, r)), max(-1.0, min(1.0, g)), max(-1.0, min(1.0, b))]
            except ValueError:
                pass
    m = re.match(
        r"^rgba?\(\s*([0-9.]+%?)\s*[, ]\s*([0-9.]+%?)\s*[, ]\s*([0-9.]+%?)(?:\s*[,/]\s*[0-9.]+%?)?\s*\)$",
        s,
    )
    parts = None
    if m:
        parts = [m.group(1), m.group(2), m.group(3)]
    else:
        toks = [t for t in re.split(r"[\s,]+", s) if t]
        if len(toks) == 3 and all(re.match(r"^-?[0-9.]+%?$", t) for t in toks):
            parts = toks
    if parts:
        def _ch(x):
            xs = str(x)
            if xs.endswith("%"):
                return max(0.0, min(100.0, float(xs[:-1]))) / 50.0 - 1.0
            n = float(xs)
            if -1.0 <= n <= 1.0 and "." in xs:
                return max(-1.0, min(1.0, n))
            return max(0.0, min(255.0, n)) / 127.5 - 1.0
        try:
            return [_ch(parts[0]), _ch(parts[1]), _ch(parts[2])]
        except (TypeError, ValueError):
            pass
    return [-1, -1, -1]

_win_color = _bgcolor_of(disp.get("bgcolor") or "#000000")
if HEADLESS:
    win = visual.Window(size=(800, 600), fullscr=False, color=_win_color,
                        units="height", allowGUI=False, winType="pyglet")
else:
    win = visual.Window(size=(int(size[0]), int(size[1])), fullscr=full,
                        color=_win_color, units="height", allowGUI=False,
                        screen=_scr_idx)

# Live keyboard runs: force en-US IME (no CJK candidate window). Autopilot/headless skip.
__IME_HELPERS__
_IME_STATE = None
_force_ime_raw = SESSION.get("force_en_ime", os.environ.get("PSYCLAW_FORCE_EN_IME", "1"))
FORCE_EN_IME = (not HEADLESS) and (
    str(_force_ime_raw).strip().lower() not in ("0", "false", "no", "off", "")
)
if FORCE_EN_IME:
    try:
        _IME_STATE = _psyclaw_force_en_ime(win)
        if _IME_STATE:
            print("[psyclaw] IME -> en-US (no CJK candidates)", flush=True)
        else:
            print("[psyclaw] WARN IME force returned empty", flush=True)
    except Exception as _ime_e:
        print("[psyclaw] WARN IME force: " + repr(_ime_e), flush=True)
    try:
        import atexit as _atexit_ime
        _atexit_ime.register(lambda: _psyclaw_restore_ime(_IME_STATE))
    except Exception:
        pass

participant_id = str(SESSION.get("participant_id") or "P01").strip() or "P01"
session_id = str(SESSION.get("session") or "1").strip() or "1"
session_date = str(SESSION.get("date") or time.strftime("%Y-%m-%d %H:%M:%S")).strip()
experimenter = str(SESSION.get("experimenter") or "").strip()
session_notes = str(SESSION.get("notes") or "").strip()
participant_name = str(SESSION.get("participant_name") or SESSION.get("name") or "").strip()
session_uid = str(SESSION.get("uid") or SESSION.get("exp_uid") or "").strip()
# user-defined session columns (from Run form extra fields)
_SESSION_RESERVED = {
    "participant_id", "session", "date", "experimenter", "notes",
    "participant_name", "name", "custom", "participant", "session_id",
    "uid", "exp_uid", "project_path",
}
SESSION_CUSTOM = {}
_raw_custom = SESSION.get("custom")
if isinstance(_raw_custom, dict):
    for _ck, _cv in _raw_custom.items():
        _k = str(_ck or "").strip()
        if _k and _k not in _SESSION_RESERVED:
            SESSION_CUSTOM[_k] = str(_cv if _cv is not None else "").strip()
for _ck, _cv in SESSION.items():
    if _ck in _SESSION_RESERVED or _ck in SESSION_CUSTOM:
        continue
    if isinstance(_cv, (str, int, float, bool)):
        SESSION_CUSTOM[str(_ck)] = str(_cv)
timestamp = time.strftime("%Y%m%d_%H%M%S")
out_csv = DATA_DIR / f"{participant_id}_s{session_id}_{timestamp}.csv"
rows = []
print(
    f"[psyclaw] design={DESIGN.get('name')!r} headless={HEADLESS} "
    f"participant={participant_id!r} name={participant_name!r} "
    f"session={session_id!r} date={session_date!r}",
    flush=True,
)

# --- instrument probe (pilot + participant): FPS + used devices ---
INSTR = {
    "headless": HEADLESS,
    "needs": {"keyboard": False, "microphone": False, "sound": False, "image": False},
    "fps_hz": None,
    "flip_ms_mean": None,
    "flip_ms_sd": None,
    "keyboard": None,
    "microphone": None,
    "sound": None,
    "display": {"size": list(size), "fullscreen": (False if HEADLESS else full), "screen": _scr_idx},
    "devices_pref": dict(DESIGN.get("devices") or {}),
    "ok": True,
    "notes": [],
}
for _r in (DESIGN.get("routines") or []):
    for _c in (_r.get("components") or []):
        _t = str((_c or {}).get("type") or "").lower()
        if _t == "keyboard":
            INSTR["needs"]["keyboard"] = True
        elif _t in ("microphone", "mic", "microphonein", "micstream"):
            INSTR["needs"]["microphone"] = True
        elif _t in ("sound", "audio"):
            INSTR["needs"]["sound"] = True
        elif _t == "image":
            INSTR["needs"]["image"] = True
if INSTR["needs"]["keyboard"]:
    _devp = DESIGN.get("devices") or {}
    _kb_dev = str(_devp.get("keyboardDevice") or "").strip()
    _mouse_dev = str(_devp.get("mouseDevice") or "").strip()
    try:
        _mouse_hz = int(_devp.get("mouseSampleRate") or 125)
    except (TypeError, ValueError):
        _mouse_hz = 125
    INSTR["keyboard"] = {
        "used": True,
        "ok": True,
        "detail": "event.getKeys path",
        "device": _kb_dev or None,
        "mouse_device": _mouse_dev or None,
        "mouse_sample_rate_hz": _mouse_hz,
    }
# FPS: prefer PsychoPy measure; fallback flip timing
try:
    _fr = win.getActualFrameRate(nIdentical=20, nMaxFrames=120, nWarmUpFrames=10, threshold=1)
    if _fr:
        INSTR["fps_hz"] = round(float(_fr), 2)
except Exception as _e:
    INSTR["notes"].append("getActualFrameRate: " + str(_e))
if INSTR["fps_hz"] is None:
    try:
        _ts = []
        for _ in range(36):
            _t0 = core.getTime()
            win.flip()
            _ts.append((core.getTime() - _t0) * 1000.0)
        _ts = _ts[6:] if len(_ts) > 6 else _ts
        if _ts:
            _mean = sum(_ts) / len(_ts)
            _var = sum((x - _mean) ** 2 for x in _ts) / max(1, len(_ts) - 1)
            INSTR["flip_ms_mean"] = round(_mean, 3)
            INSTR["flip_ms_sd"] = round(_var ** 0.5, 3)
            if _mean > 0:
                INSTR["fps_hz"] = round(1000.0 / _mean, 2)
    except Exception as _e2:
        INSTR["notes"].append("flip_probe: " + str(_e2))
if INSTR["fps_hz"] is not None and INSTR["fps_hz"] < 20:
    INSTR["ok"] = False
    INSTR["notes"].append("low FPS < 20 Hz")
# Microphone only if design uses it
if INSTR["needs"]["microphone"]:
    _devp = DESIGN.get("devices") or {}
    _mic_dev = str(_devp.get("micDevice") or _devp.get("micLabel") or "").strip()
    try:
        _mic_hz = int(_devp.get("micSampleRate") or 44100)
    except (TypeError, ValueError):
        _mic_hz = 44100
    try:
        try:
            from psychopy import microphone as _micmod  # type: ignore
            INSTR["microphone"] = {
                "used": True, "ok": True,
                "detail": "psychopy.microphone import ok",
                "device": _mic_dev or None,
                "sample_rate_hz": _mic_hz,
            }
        except Exception:
            from psychopy.sound import microphone as _sm  # type: ignore
            INSTR["microphone"] = {
                "used": True, "ok": True,
                "detail": "psychopy.sound.microphone import ok",
                "device": _mic_dev or None,
                "sample_rate_hz": _mic_hz,
            }
    except Exception as _em:
        INSTR["microphone"] = {
            "used": True, "ok": False, "detail": str(_em),
            "device": _mic_dev or None, "sample_rate_hz": _mic_hz,
        }
        INSTR["ok"] = False
        INSTR["notes"].append("microphone unavailable")
else:
    INSTR["microphone"] = {"used": False, "ok": True, "detail": "not used in design"}
if INSTR["needs"]["sound"]:
    _devp2 = DESIGN.get("devices") or {}
    _spk_dev = str(_devp2.get("speakerDevice") or _devp2.get("speakerLabel") or "").strip()
    try:
        _spk_hz = int(_devp2.get("speakerSampleRate") or 44100)
    except (TypeError, ValueError):
        _spk_hz = 44100
    try:
        from psychopy import sound as _snd  # noqa: F401
        INSTR["sound"] = {
            "used": True, "ok": True, "detail": "psychopy.sound import ok",
            "device": _spk_dev or None, "sample_rate_hz": _spk_hz,
        }
    except Exception as _es:
        INSTR["sound"] = {
            "used": True, "ok": False, "detail": str(_es),
            "device": _spk_dev or None, "sample_rate_hz": _spk_hz,
        }
        INSTR["ok"] = False
else:
    INSTR["sound"] = {"used": False, "ok": True, "detail": "not used in design"}
print("[psyclaw] INSTR " + json.dumps(INSTR, ensure_ascii=False), flush=True)

NOT_STARTED = 0
STARTED = 1
FINISHED = 2

def color_of(c):
    """Component color: name / hex / rgb → PsychoPy [-1,1]. Unknown → white."""
    if isinstance(c, (list, tuple)) and len(c) >= 3:
        return list(c)[:3]
    s = str(c or "white").strip()
    if not s or s.startswith("$"):
        return [1, 1, 1]
    low = s.lower().strip()
    named_ok = low in {
        "black", "white", "gray", "grey",
        "darkgray", "darkgrey", "dark gray", "dark grey",
        "lightgray", "lightgrey", "light gray", "light grey",
        "red", "green", "blue", "yellow", "cyan", "magenta", "orange",
        "lime", "aqua", "fuchsia",
    }
    looks_hex = low.startswith("#") and len(low) in (4, 7)
    looks_rgb = low.startswith("rgb")
    parts = [x for x in re.split(r"[\s,]+", low) if x]
    looks_triplet = len(parts) == 3 and all(re.match(r"^-?[0-9.]+%?$", p) for p in parts)
    if named_ok or looks_hex or looks_rgb or looks_triplet:
        return _bgcolor_of(s)
    return [1, 1, 1]

def _load_conditions_rows(node):
    """Return list[dict] for loop stimlist, or None if pure nReps."""
    emb = node.get("conditions")
    if isinstance(emb, list) and emb:
        out = []
        for row in emb:
            if isinstance(row, dict):
                out.append(dict(row))
        return out or None
    path = node.get("conditionsFile") or node.get("conditions_file") or ""
    path = str(path).strip()
    if not path:
        return None
    p = _Path(path)
    if not p.is_file():
        # try next to script / data
        for base in (_Path("."), DATA_DIR, _Path(__file__).resolve().parent):
            cand = base / path
            if cand.is_file():
                p = cand
                break
    if not p.is_file():
        print(f"[psyclaw] WARN conditionsFile not found: {path!r}", flush=True)
        return None
    suf = p.suffix.lower()
    try:
        if suf in (".csv", ".txt"):
            with open(p, "r", encoding="utf-8-sig", newline="") as f:
                rdr = _csv.DictReader(f)
                return [dict(r) for r in rdr]
        if suf in (".xlsx", ".xlsm", ".xls"):
            try:
                import openpyxl
            except ImportError:
                print("[psyclaw] WARN openpyxl missing; cannot read xlsx", flush=True)
                return None
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            headers = next(it, None)
            if not headers:
                return None
            headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(headers)]
            out = []
            for row in it:
                d = {}
                empty = True
                for i, h in enumerate(headers):
                    v = row[i] if i < len(row) else None
                    if v is not None and str(v).strip() != "":
                        empty = False
                    d[h] = "" if v is None else v
                if not empty:
                    out.append(d)
            return out or None
    except Exception as exc:
        print(f"[psyclaw] WARN load conditions failed: {exc!r}", flush=True)
        return None
    return None

def resolve_val(val, trial_vars):
    """PsychoPy-like $var substitution in string params."""
    if not isinstance(val, str):
        return val
    tv = trial_vars or {}
    s = val.strip()
    if s.startswith("$") and s[1:] in tv:
        return tv[s[1:]]
    if "$" not in val:
        return val
    def repl(m):
        k = m.group(1)
        return str(tv[k]) if k in tv else m.group(0)
    return re.sub(r"\$([A-Za-z_][A-Za-z0-9_]*)", repl, val)

def resolve_params(params, trial_vars):
    out = {}
    for k, v in (params or {}).items():
        out[k] = resolve_val(v, trial_vars)
    return out

def expand_flow(flow, loop_name="", this_n=None, n_reps=None, trial_vars=None):
    """Expand flow tree into flat trial steps.

    Loop without conditions: nReps times full children sequence.
    Loop with conditions (embedded or file): nReps × condition rows.
      loopType: sequential | random (shuffle each rep) | fullRandom (shuffle all).
    Nested loops recurse. trial_vars merge outer→inner.
    """
    out = []
    base_vars = dict(trial_vars or {})
    for node in flow or []:
        kind = node.get("kind")
        if kind == "loop":
            n = max(1, int(node.get("nReps") or 1))
            kids = node.get("children") or []
            lname = node.get("name") or "loop"
            loop_type = str(node.get("loopType") or node.get("loop_type") or "sequential").lower().replace("_", "")
            if loop_type in ("fullrandom", "full-random"):
                loop_type = "fullrandom"
            conds = _load_conditions_rows(node)
            if not conds:
                for rep in range(n):
                    out.extend(expand_flow(
                        kids, loop_name=lname, this_n=rep, n_reps=n, trial_vars=base_vars
                    ))
            else:
                schedule = []  # list of (thisN_label, cond_dict)
                if loop_type == "fullrandom":
                    for rep in range(n):
                        for ci, cond in enumerate(conds):
                            schedule.append((rep * len(conds) + ci, dict(cond)))
                    random.shuffle(schedule)
                    # re-index thisN after shuffle
                    schedule = [(i, c) for i, (_, c) in enumerate(schedule)]
                else:
                    idx = 0
                    for rep in range(n):
                        order = list(range(len(conds)))
                        if loop_type == "random":
                            random.shuffle(order)
                        for ci in order:
                            schedule.append((idx, dict(conds[ci])))
                            idx += 1
                total = len(schedule)
                for this_i, cond in schedule:
                    merged = dict(base_vars)
                    merged.update(cond)
                    out.extend(expand_flow(
                        kids, loop_name=lname, this_n=this_i, n_reps=total, trial_vars=merged
                    ))
        elif kind == "routine":
            out.append({
                "routine": node.get("routine"),
                "loop": loop_name or "",
                "thisN": this_n if this_n is not None else "",
                "nReps": n_reps if n_reps is not None else "",
                "trialVars": dict(base_vars),
            })
    return out

routines = {r["name"]: r for r in DESIGN.get("routines") or []}
sequence = expand_flow(DESIGN.get("flow") or [])
print(f"[psyclaw] expanded {len(sequence)} steps", flush=True)
if HEADLESS and len(sequence) > 12:
    seq2 = []
    for step in sequence:
        name = step.get("routine")
        trials = [x for x in seq2 if x.get("routine") not in ("instructions", "thanks")]
        if name in ("instructions", "thanks") or len(trials) < 4:
            seq2.append(step)
    if any(s.get("routine") == "thanks" for s in sequence):
        if not seq2 or seq2[-1].get("routine") != "thanks":
            seq2.append({"routine": "thanks", "loop": "", "thisN": "", "nReps": "", "trialVars": {}})
    sequence = seq2
    print(f"[psyclaw] headless truncated to {len(sequence)} steps", flush=True)

global_clock = core.Clock()
trial_i = 0
end_exp = False
# ESC (or future user abort keys) → manual end, not clean finished
abort_reason = None

for step in sequence:
    if end_exp:
        break
    if FORCE_EN_IME:
        try:
            _psyclaw_assert_en_ime()
        except Exception:
            pass
    rname = step.get("routine")
    r = routines.get(rname)
    if not r:
        continue

    trial_vars = step.get("trialVars") or {}
    comps = r.get("components") or []
    stims = []
    max_end = 0.0
    has_open = False

    for c in comps:
        start = float(c.get("start") or 0)
        dur = c.get("duration")
        try:
            open_ended = dur is None or dur == "" or float(dur) == -1
        except (TypeError, ValueError):
            open_ended = True
        if open_ended:
            has_open = True
            end = None
        else:
            end = start + float(dur)
            max_end = max(max_end, end)
        typ = c.get("type")
        params = resolve_params(c.get("params") or {}, trial_vars)
        obj = None
        if typ in ("text", "fixation"):
            obj = visual.TextStim(
                win,
                text=str(params.get("text", "+" if typ == "fixation" else "")),
                height=float(params.get("height") or (0.08 if typ == "fixation" else 0.05)),
                color=color_of(params.get("color", "white")),
            )
        elif typ == "image":
            try:
                obj = visual.ImageStim(
                    win, image=str(params.get("path") or "missing.png"),
                    size=float(params.get("size") or 0.5),
                )
            except Exception:
                obj = visual.Rect(win, width=0.3, height=0.3, fillColor=[0.2, 0.2, 0.2])
        elif typ == "video":
            # PsychoPy MovieStim (Movie component). Fail soft -> gray rect.
            try:
                _vpath = str(params.get("path") or "missing.mp4")
                _vsz = float(params.get("size") or 0.5)
                try:
                    _vvol = float(params.get("volume")) if params.get("volume") is not None else 1.0
                except (TypeError, ValueError):
                    _vvol = 1.0
                _vvol = max(0.0, min(1.0, _vvol))
                try:
                    obj = visual.MovieStim(
                        win, filename=_vpath, size=_vsz, units="height",
                        volume=_vvol, loop=False, autoStart=False,
                    )
                except TypeError:
                    try:
                        obj = visual.MovieStim(
                            win, filename=_vpath, size=_vsz, units="height", loop=False,
                        )
                    except Exception:
                        obj = visual.MovieStim(win, filename=_vpath, size=_vsz)
            except Exception:
                obj = visual.Rect(win, width=0.45, height=0.3, fillColor=[0.12, 0.12, 0.18])
        stims.append({
            "obj": obj,
            "start": start,
            "end": end,
            "open": open_ended,
            "typ": typ,
            "params": params,
            "name": c.get("name") or typ,
            "status": NOT_STARTED,
        })

    if not has_open and max_end <= 0:
        max_end = 0.5

    routine_clock = core.Clock()
    continue_routine = True
    resp = None
    rt = None
    resp_keys = ""
    force_end = False
    headless_fire_at = None

    event.clearEvents(eventType="keyboard")

    while continue_routine:
        t = routine_clock.getTime()

        for s in stims:
            if s["status"] == NOT_STARTED and t >= s["start"]:
                s["status"] = STARTED
                if s["typ"] == "keyboard":
                    event.clearEvents(eventType="keyboard")
                    if HEADLESS:
                        headless_fire_at = t + 0.05
                if s["typ"] == "video" and s["obj"] is not None:
                    try:
                        if hasattr(s["obj"], "seek"):
                            try:
                                s["obj"].seek(0.0)
                            except Exception:
                                pass
                        if hasattr(s["obj"], "play"):
                            s["obj"].play()
                    except Exception:
                        pass
            if s["status"] == STARTED and not s["open"] and s["end"] is not None and t >= s["end"]:
                s["status"] = FINISHED
                if s["typ"] == "video" and s["obj"] is not None:
                    try:
                        if hasattr(s["obj"], "pause"):
                            s["obj"].pause()
                        elif hasattr(s["obj"], "stop"):
                            s["obj"].stop()
                    except Exception:
                        pass

        for s in stims:
            if s["status"] == STARTED and s["obj"] is not None:
                s["obj"].draw()
        win.flip()

        for s in stims:
            if s["typ"] != "keyboard" or s["status"] != STARTED:
                continue
            keys_str = str(s["params"].get("keys") or "space")
            keylist = [k.strip() for k in keys_str.replace("+", ",").split(",") if k.strip()]
            if HEADLESS:
                if headless_fire_at is not None and t >= headless_fire_at:
                    _kind = detect_trial_kind(trial_vars)
                    resp_keys = keys_str
                    if _kind == "nogo":
                        # correct withhold — do not inject a key (FA would break rates)
                        resp = None
                        rt = None
                        force_end = False
                        s["status"] = FINISHED
                        headless_fire_at = None
                        if s.get("open"):
                            continue_routine = False
                    else:
                        # prefer corrAns / correct from conditions when present
                        corr = trial_vars.get("corrAns") or trial_vars.get("correct") or trial_vars.get("corr")
                        if corr is not None and str(corr).strip() != "":
                            resp = str(corr).strip().lower()
                        else:
                            resp = keylist[0] if keylist else "space"
                        rt = round(0.2 + 0.1 * (trial_i % 5), 4)
                        force_end = bool(s["params"].get("force_end", True))
                        s["status"] = FINISHED
                        headless_fire_at = None
                        if force_end:
                            continue_routine = False
                break
            got = event.getKeys(keyList=keylist, timeStamped=routine_clock)
            if got:
                resp = got[0][0]
                rt = round(got[0][1] - s["start"], 4)
                resp_keys = keys_str
                force_end = bool(s["params"].get("force_end", True))
                s["status"] = FINISHED
                if force_end:
                    continue_routine = False
                break

        if not continue_routine:
            break

        if not has_open and t >= max_end:
            continue_routine = False
            break

        if HEADLESS and has_open and t > 2.0:
            continue_routine = False
            break

        if not HEADLESS:
            esc = event.getKeys(keyList=["escape"])
            if esc:
                end_exp = True
                abort_reason = "escape"
                print("[psyclaw] user abort: escape", flush=True)
                continue_routine = False
                break

        if stims and all(s["status"] == FINISHED or (s["open"] and s["status"] != STARTED) for s in stims):
            pending_open = any(s["open"] and s["status"] == NOT_STARTED for s in stims)
            if not pending_open and not any(s["status"] == STARTED for s in stims):
                if not has_open:
                    continue_routine = False

    trial_i += 1
    _had_kb = any(s.get("typ") == "keyboard" for s in stims)
    row = {
        "trial": trial_i,
        "routine": rname,
        "loop": step.get("loop") or "",
        "thisN": step.get("thisN") if step.get("thisN") != "" else "",
        "nReps": step.get("nReps") if step.get("nReps") != "" else "",
        "response": resp or "",
        "rt": rt if rt is not None else "",
        "keys": resp_keys,
        "participant_id": participant_id,
        "participant_name": participant_name,
        "session": session_id,
        "session_date": session_date,
        "uid": session_uid,
        "experimenter": experimenter,
        "notes": session_notes,
        "t_global": round(global_clock.getTime(), 4),
    }
    row.update(SESSION_CUSTOM)
    for k, v in trial_vars.items():
        if k not in row:
            row[k] = v
    # corrAns + corr (0/1) — after stimlist merge so computed corr wins
    apply_trial_scores(row, resp=resp, trial_vars=trial_vars, had_keyboard=_had_kb)
    rows.append(row)
    print(
        f"[psyclaw] {trial_i} routine={rname} loop={step.get('loop')!r} "
        f"thisN={step.get('thisN')} vars={list(trial_vars.keys())} "
        f"resp={resp} rt={rt} corr={row.get('corr')!r}",
        flush=True,
    )

if rows:
    # preferred column order, then remaining keys (condition cols etc.)
    _pref = [
        "trial", "routine", "loop", "thisN", "nReps",
        "response", "corrAns", "corr", "rt", "keys",
        "participant_id", "participant_name", "session", "session_date", "uid",
        "experimenter", "notes", "t_global",
    ]
    fieldnames = []
    seen = set()
    for k in _pref:
        fieldnames.append(k)
        seen.add(k)
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                fieldnames.append(k)
else:
    # still write header-only CSV (PsychoPy-desktop-like empty data file)
    fieldnames = [
        "trial", "routine", "loop", "thisN", "nReps",
        "response", "corrAns", "corr", "rt", "keys",
        "participant_id", "participant_name", "session", "session_date", "uid",
        "experimenter", "notes", "t_global",
    ]
    for _ck in SESSION_CUSTOM.keys():
        if _ck not in fieldnames:
            fieldnames.append(_ck)
with open(out_csv, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
    w.writeheader()
    if rows:
        w.writerows(rows)
print(f"[psyclaw] wrote {len(rows)} rows -> {out_csv}", flush=True)

# Run-level metrics: accuracy / mean RT / optional group_by (design.metrics)
_metrics_cfg = DESIGN.get("metrics") if isinstance(DESIGN.get("metrics"), dict) else {}
_summary = summarize_rows(rows, _metrics_cfg)
_summary["participant_id"] = participant_id
_summary["session"] = session_id
_summary["uid"] = session_uid
_summary["design"] = str(DESIGN.get("name") or "")
_meta_ids = {
    "participant_id": participant_id,
    "session": session_id,
    "uid": session_uid,
    "design": str(DESIGN.get("name") or ""),
}
_by_cond_rows = by_condition_rows(_summary, **_meta_ids)
_long_rows = metrics_long_rows(_summary, **_meta_ids)
_summary_name = out_csv.stem + "_summary.json"
_by_cond_name = out_csv.stem + "_by_condition.csv"
_long_name = out_csv.stem + "_metrics_long.csv"
out_summary = DATA_DIR / _summary_name
out_by_cond = DATA_DIR / _by_cond_name
out_long = DATA_DIR / _long_name
try:
    out_summary.write_text(json.dumps(_summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[psyclaw] wrote metrics summary -> {out_summary}", flush=True)
    _ov = _summary.get("overall") or {}
    print(
        f"[psyclaw] metrics overall accuracy={_ov.get('accuracy')!r} "
        f"mean_rt={_ov.get('mean_rt')!r} n_scored={_ov.get('n_scored')!r}",
        flush=True,
    )
except Exception as _se:
    print(f"[psyclaw] WARN metrics summary: {_se}", flush=True)
    out_summary = None

def _write_metric_csv(path, rowlist):
    if not rowlist:
        # still write header shell from preferred keys
        prefs = [
            "participant_id", "session", "uid", "design", "scope", "group_key",
            "metric", "value", "n", "n_scored", "accuracy", "mean_rt",
        ]
        with open(path, "w", newline="", encoding="utf-8") as _mf:
            csv.DictWriter(_mf, fieldnames=prefs).writeheader()
        return
    fn = []
    seen = set()
    for _r in rowlist:
        for _k in _r.keys():
            if _k not in seen:
                seen.add(_k)
                fn.append(_k)
    with open(path, "w", newline="", encoding="utf-8") as _mf:
        _w = csv.DictWriter(_mf, fieldnames=fn, extrasaction="ignore")
        _w.writeheader()
        _w.writerows(rowlist)

try:
    _write_metric_csv(out_by_cond, _by_cond_rows)
    print(f"[psyclaw] wrote by-condition -> {out_by_cond}", flush=True)
except Exception as _be:
    print(f"[psyclaw] WARN by_condition: {_be}", flush=True)
    out_by_cond = None
try:
    _write_metric_csv(out_long, _long_rows)
    print(f"[psyclaw] wrote metrics long -> {out_long}", flush=True)
except Exception as _le:
    print(f"[psyclaw] WARN metrics_long: {_le}", flush=True)
    out_long = None

# Mirror into project/data/ like PsychoPy desktop (next to experiment folder)
_proj_csv = None
_proj_summary = None
_proj_by_cond = None
_proj_long = None
_proj = str(SESSION.get("project_path") or "").strip()
if _proj:
    try:
        _pdata = _Path(_proj) / "data"
        _pdata.mkdir(parents=True, exist_ok=True)

        def _mirror_one(src):
            if src is None:
                return None
            try:
                p = _Path(src)
                if not p.exists():
                    return None
                dest = _pdata / p.name
                if p.resolve() == dest.resolve():
                    return dest
                _shutil.copy2(p, dest)
                return dest
            except Exception as _me1:
                print(f"[psyclaw] WARN mirror {src}: {_me1}", flush=True)
                return None

        _proj_csv = _mirror_one(out_csv)
        if _proj_csv is not None:
            print(f"[psyclaw] mirrored CSV -> {_proj_csv}", flush=True)
        _proj_summary = _mirror_one(out_summary)
        if _proj_summary is not None:
            print(f"[psyclaw] mirrored summary -> {_proj_summary}", flush=True)
        _proj_by_cond = _mirror_one(out_by_cond)
        if _proj_by_cond is not None:
            print(f"[psyclaw] mirrored by_condition -> {_proj_by_cond}", flush=True)
        _proj_long = _mirror_one(out_long)
        if _proj_long is not None:
            print(f"[psyclaw] mirrored metrics_long -> {_proj_long}", flush=True)
    except Exception as _me:
        print(f"[psyclaw] WARN project data mirror: {_me}", flush=True)
try:
    INSTR["n_rows"] = len(rows)
    INSTR["csv"] = str(out_csv.name)
    if _proj_csv is not None:
        INSTR["csv_project"] = str(_proj_csv)
    INSTR["metrics"] = _summary
    if out_summary is not None:
        INSTR["summary"] = str(out_summary.name if hasattr(out_summary, "name") else out_summary)
    if _proj_summary is not None:
        INSTR["summary_project"] = str(_proj_summary)
    if out_by_cond is not None:
        INSTR["by_condition"] = str(out_by_cond.name if hasattr(out_by_cond, "name") else out_by_cond)
    if _proj_by_cond is not None:
        INSTR["by_condition_project"] = str(_proj_by_cond)
    if out_long is not None:
        INSTR["metrics_long"] = str(out_long.name if hasattr(out_long, "name") else out_long)
    if _proj_long is not None:
        INSTR["metrics_long_project"] = str(_proj_long)
except Exception:
    pass

try:
    if HEADLESS or participant_id.startswith("P_autopilot"):
        INSTR["mode"] = "autopilot"
    elif participant_id.startswith("P_pilot"):
        INSTR["mode"] = "pilot"
    else:
        INSTR["mode"] = "participant"
    INSTR["design_name"] = str(DESIGN.get("name") or "")
    INSTR["session"] = {
        "participant_id": participant_id,
        "participant_name": participant_name,
        "session": session_id,
        "date": session_date,
        "uid": session_uid,
        "experimenter": experimenter,
        "notes": session_notes,
        "custom": dict(SESSION_CUSTOM) if SESSION_CUSTOM else {},
    }
    INSTR["at"] = time.strftime("%Y-%m-%d %H:%M:%S")
    # end_status: normal | manual (ESC / user abort) — UI roster chip
    if abort_reason:
        INSTR["end_status"] = "manual"
        INSTR["end_reason"] = str(abort_reason)
    else:
        INSTR["end_status"] = "normal"
        INSTR["end_reason"] = ""
except Exception:
    pass
_end_payload = {
    "end_status": "manual" if abort_reason else "normal",
    "reason": str(abort_reason or ""),
    "n_rows": len(rows),
}
try:
    (DATA_DIR / "end_reason.json").write_text(
        json.dumps(_end_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        "[psyclaw] END "
        + _end_payload["end_status"]
        + ((" reason=" + _end_payload["reason"]) if _end_payload["reason"] else ""),
        flush=True,
    )
except Exception as _er:
    print("[psyclaw] WARN end_reason.json: " + str(_er), flush=True)
try:
    (DATA_DIR / "instrument.json").write_text(
        json.dumps(INSTR, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print("[psyclaw] wrote instrument.json", flush=True)
except Exception as _iw:
    print("[psyclaw] WARN instrument.json: " + str(_iw), flush=True)
try:
    if FORCE_EN_IME:
        _psyclaw_restore_ime(_IME_STATE)
        _IME_STATE = None
except Exception:
    pass
try:
    win.close()
except Exception:
    pass
# ESC abort: non-zero exit so harness maps to stopped + end_status=manual
# (exit 0 would look like clean finished). CSV/instrument already written.
if abort_reason:
    import sys as _sys
    _sys.exit(130)
core.quit()
'''


def _normalize_loop_type(raw: Any) -> str:
    t = str(raw or "sequential").lower().replace("_", "").replace("-", "")
    if t in ("fullrandom",):
        return "fullrandom"
    if t in ("random",):
        return "random"
    return "sequential"


def expand_flow_py(
    flow: Optional[List[Dict[str, Any]]],
    loop_name: str = "",
    this_n=None,
    n_reps=None,
    trial_vars: Optional[Dict[str, Any]] = None,
    rng: Optional[random.Random] = None,
) -> List[Dict[str, Any]]:
    """Host-side expand (tests / tooling). Same semantics as runtime expand_flow."""
    rng = rng or random.Random(0)  # deterministic default for tests
    out: List[Dict[str, Any]] = []
    base_vars = dict(trial_vars or {})
    for node in flow or []:
        kind = node.get("kind")
        if kind == "loop":
            n = max(1, int(node.get("nReps") or 1))
            kids = node.get("children") or []
            lname = node.get("name") or "loop"
            loop_type = _normalize_loop_type(node.get("loopType") or node.get("loop_type"))
            conds = node.get("conditions")
            if not isinstance(conds, list) or not conds:
                for rep in range(n):
                    out.extend(
                        expand_flow_py(
                            kids, loop_name=lname, this_n=rep, n_reps=n, trial_vars=base_vars, rng=rng
                        )
                    )
            else:
                schedule: List[tuple] = []
                if loop_type == "fullrandom":
                    for rep in range(n):
                        for ci, cond in enumerate(conds):
                            if isinstance(cond, dict):
                                schedule.append((rep * len(conds) + ci, dict(cond)))
                    rng.shuffle(schedule)
                    schedule = [(i, c) for i, (_, c) in enumerate(schedule)]
                else:
                    idx = 0
                    for rep in range(n):
                        order = list(range(len(conds)))
                        if loop_type == "random":
                            rng.shuffle(order)
                        for ci in order:
                            cond = conds[ci]
                            if isinstance(cond, dict):
                                schedule.append((idx, dict(cond)))
                                idx += 1
                total = len(schedule)
                for this_i, cond in schedule:
                    merged = dict(base_vars)
                    merged.update(cond)
                    out.extend(
                        expand_flow_py(
                            kids,
                            loop_name=lname,
                            this_n=this_i,
                            n_reps=total,
                            trial_vars=merged,
                            rng=rng,
                        )
                    )
        elif kind == "routine":
            out.append(
                {
                    "routine": node.get("routine"),
                    "loop": loop_name or "",
                    "thisN": this_n if this_n is not None else "",
                    "nReps": n_reps if n_reps is not None else "",
                    "trialVars": dict(base_vars),
                }
            )
    return out


def _trial_metrics_inline_src() -> str:
    """Load trial_metrics.py for inlining into generated experiment.py.

    Strip ``from __future__`` (must be file-top only) so mid-script paste is valid.
    """
    path = Path(__file__).resolve().with_name("trial_metrics.py")
    text = path.read_text(encoding="utf-8")
    lines = [ln for ln in text.splitlines() if not ln.startswith("from __future__")]
    return "\n".join(lines).rstrip() + "\n"


def compile_design(
    design: Dict[str, Any],
    session: Optional[Dict[str, Any]] = None,
) -> str:
    """Compile design.json → experiment.py. Session comes from Run form (not PsychoPy Dlg)."""
    payload = json.dumps(design, ensure_ascii=False)
    sess = dict(session or {})
    # normalize keys used in runner
    if "participant_id" not in sess and sess.get("participant"):
        sess["participant_id"] = sess.get("participant")
    sess_payload = json.dumps(sess, ensure_ascii=False)
    try:
        from ime_guard import EXPERIMENT_IME_HELPERS
    except ImportError:
        from backend.ime_guard import EXPERIMENT_IME_HELPERS  # type: ignore
    body = (
        _RUNNER
        .replace("__TRIAL_METRICS_SRC__", _trial_metrics_inline_src())
        .replace("__IME_HELPERS__", EXPERIMENT_IME_HELPERS)
        .replace("__DESIGN_JSON__", json.dumps(payload))
        .replace("__SESSION_JSON__", json.dumps(sess_payload))
    )
    return HEADER + body


def compile_any(
    *,
    paradigm_id: str = "",
    spec: Optional[Dict[str, Any]] = None,
    design: Optional[Dict[str, Any]] = None,
    session: Optional[Dict[str, Any]] = None,
) -> str:
    if design and isinstance(design, dict) and design.get("routines"):
        # prefer explicit session; else pull from spec.session / spec fields
        sess = session
        if sess is None and isinstance(spec, dict):
            if isinstance(spec.get("session"), dict):
                sess = spec.get("session")
            else:
                sess = {
                                    "participant_id": spec.get("participant_id") or "P01",
                                    "session": spec.get("session_id") or spec.get("session") or "1",
                                    "date": spec.get("date") or "",
                                    "experimenter": spec.get("experimenter") or "",
                                    "notes": spec.get("notes") or "",
                                    "participant_name": spec.get("participant_name") or "",
                                }
        sess = dict(sess or {})
        # project_path → SESSION so runner can mirror CSV into project/data/
        if isinstance(spec, dict):
            pp = str(spec.get("project_path") or "").strip()
            if pp:
                sess["project_path"] = pp
            if isinstance(spec.get("session"), dict):
                pp2 = str(spec["session"].get("project_path") or "").strip()
                if pp2 and not sess.get("project_path"):
                    sess["project_path"] = pp2
        return compile_design(design, session=sess)
    return compile_spec(paradigm_id or "stroop", spec or {})


def parse_conditions_bytes(filename: str, data: bytes) -> Dict[str, Any]:
    """Parse csv/xlsx bytes → {filename, columns, rows, n}."""
    name = filename or "conditions.csv"
    suf = Path(name).suffix.lower()
    rows: List[Dict[str, Any]] = []
    if suf in (".csv", ".txt", ""):
        text = data.decode("utf-8-sig", errors="replace")
        rdr = csv.DictReader(text.splitlines())
        rows = [dict(r) for r in rdr]
    elif suf in (".xlsx", ".xlsm", ".xls"):
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = next(it, None)
        if not headers:
            return {"filename": name, "columns": [], "rows": [], "n": 0}
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(headers)]
        for row in it:
            d: Dict[str, Any] = {}
            empty = True
            for i, h in enumerate(headers):
                v = row[i] if i < len(row) else None
                if v is not None and str(v).strip() != "":
                    empty = False
                d[h] = "" if v is None else v
            if not empty:
                rows.append(d)
    else:
        raise ValueError(f"unsupported conditions file type: {suf or '(none)'}")
    cols: List[str] = []
    seen = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                cols.append(k)
    return {"filename": name, "columns": cols, "rows": rows, "n": len(rows)}
